from rest_framework import viewsets, status
from rest_framework.decorators import action, api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.parsers import MultiPartParser
from django.contrib.auth import get_user_model
from django.contrib.auth.hashers import make_password, check_password
from django.contrib.auth.password_validation import validate_password
from django.core.exceptions import ValidationError as DjangoValidationError
from django.http import HttpResponse
from django.utils.text import slugify
from django.utils import timezone
from django.db import transaction
from datetime import time as time_type, timedelta
import secrets
import openpyxl
import io

from rest_framework_simplejwt.views import TokenObtainPairView
from rest_framework_simplejwt.serializers import TokenObtainPairSerializer
from rest_framework_simplejwt.exceptions import AuthenticationFailed

from .models import Institution, Department, Classroom, Course, Schedule, Unavailability, StudentEnrollment, Announcement, UserNotification, AttendanceSession, AttendanceRecord, ChatMessage, CourseMaterial, Grade, AuditLog
from .serializers import (
    InstitutionSerializer, UserSerializer, DepartmentSerializer,
    ClassroomSerializer, CourseSerializer, ScheduleSerializer,
    StudentEnrollmentSerializer, AnnouncementSerializer, UserNotificationSerializer,
    AttendanceSessionSerializer, AttendanceRecordSerializer,
    ChatMessageSerializer, CourseMaterialSerializer, GradeSerializer,
)

User = get_user_model()


def _get_client_ip(request):
    forwarded = request.META.get('HTTP_X_FORWARDED_FOR')
    if forwarded:
        return forwarded.split(',')[0].strip()
    return request.META.get('REMOTE_ADDR')


def audit(request_or_actor, action: str, model_name: str = '', object_id='', description: str = '', request=None):
    actor = None
    ip = None
    if request is not None:
        ip = _get_client_ip(request)
        actor = request.user if request.user.is_authenticated else None
    elif hasattr(request_or_actor, 'is_authenticated'):
        actor = request_or_actor if request_or_actor.is_authenticated else None
    AuditLog.objects.create(
        actor=actor,
        action=action,
        model_name=model_name,
        object_id=str(object_id),
        description=description,
        ip_address=ip,
    )


def _t(request, tr_msg: str, en_msg: str = "") -> str:
    """Return Turkish or English message based on Accept-Language header."""
    lang = request.META.get("HTTP_ACCEPT_LANGUAGE", "tr")
    if lang.startswith("en"):
        return en_msg if en_msg else tr_msg
    return tr_msg


# ─────────────────────────── Yardımcı Fonksiyonlar ──────────────────────────

TITLE_PREFIXES = [
    'Prof. Dr.', 'Doç. Dr.', 'Dr. Öğr. Üyesi', 'Dr.',
    'Öğr. Gör. Dr.', 'Öğr. Gör.',
    'Arş. Gör. Dr.', 'Arş. Gör.',
    'Assoc. Prof. Dr.', 'Assoc. Prof.',
    'Assist. Prof. Dr.', 'Assist. Prof.',
    'Prof.',
]

DAY_NORMALIZE = {
    'mon': 'Monday', 'monday': 'Monday', 'pazartesi': 'Monday',
    'tue': 'Tuesday', 'tuesday': 'Tuesday', 'sali': 'Tuesday', 'salı': 'Tuesday',
    'wed': 'Wednesday', 'wednesday': 'Wednesday', 'carsamba': 'Wednesday', 'çarşamba': 'Wednesday',
    'thu': 'Thursday', 'thursday': 'Thursday', 'persembe': 'Thursday', 'perşembe': 'Thursday',
    'fri': 'Friday', 'friday': 'Friday', 'cuma': 'Friday',
}

ALL_SLOTS = [
    ('Monday',    time_type(8,  0),  time_type(10, 0)),
    ('Monday',    time_type(10, 0),  time_type(12, 0)),
    ('Monday',    time_type(13, 0),  time_type(15, 0)),
    ('Monday',    time_type(15, 0),  time_type(17, 0)),
    ('Tuesday',   time_type(8,  0),  time_type(10, 0)),
    ('Tuesday',   time_type(10, 0),  time_type(12, 0)),
    ('Tuesday',   time_type(13, 0),  time_type(15, 0)),
    ('Tuesday',   time_type(15, 0),  time_type(17, 0)),
    ('Wednesday', time_type(8,  0),  time_type(10, 0)),
    ('Wednesday', time_type(10, 0),  time_type(12, 0)),
    ('Wednesday', time_type(13, 0),  time_type(15, 0)),
    ('Wednesday', time_type(15, 0),  time_type(17, 0)),
    ('Thursday',  time_type(8,  0),  time_type(10, 0)),
    ('Thursday',  time_type(10, 0),  time_type(12, 0)),
    ('Thursday',  time_type(13, 0),  time_type(15, 0)),
    ('Thursday',  time_type(15, 0),  time_type(17, 0)),
    ('Friday',    time_type(8,  0),  time_type(10, 0)),
    ('Friday',    time_type(10, 0),  time_type(12, 0)),
    ('Friday',    time_type(13, 0),  time_type(15, 0)),
    ('Friday',    time_type(15, 0),  time_type(17, 0)),
]


def normalize_day(day: str) -> str:
    return DAY_NORMALIZE.get(day.lower().strip(), day)


def normalize_turkish(text: str) -> str:
    return (text
            .replace('ğ', 'g').replace('Ğ', 'G')
            .replace('ü', 'u').replace('Ü', 'U')
            .replace('ş', 's').replace('Ş', 'S')
            .replace('ı', 'i').replace('İ', 'I')
            .replace('ö', 'o').replace('Ö', 'O')
            .replace('ç', 'c').replace('Ç', 'C'))


def extract_title_and_name(raw: str):
    """'Prof. Dr. Ali Yılmaz' → ('Prof. Dr.', 'Ali Yılmaz')"""
    raw = raw.strip()
    for t in sorted(TITLE_PREFIXES, key=len, reverse=True):
        if raw.startswith(t):
            return t, raw[len(t):].strip()
    return '', raw


def make_username(first: str, last: str) -> str:
    f = normalize_turkish(first.lower()).strip().replace(' ', '_')
    l = normalize_turkish(last.lower()).strip().replace(' ', '_')
    base = f"{f}_{l}" if l else f
    base = ''.join(c if c.isalnum() or c == '_' else '' for c in base)
    # Çakışma varsa sayı ekle
    username = base
    counter = 1
    while User.objects.filter(username=username).exists():
        username = f"{base}{counter}"
        counter += 1
    return username


def gen_password(length=12) -> str:
    chars = 'ABCDEFGHJKLMNPQRSTUVWXYZabcdefghjkmnpqrstuvwxyz23456789!@#$%'
    return ''.join(secrets.choice(chars) for _ in range(length))


def has_schedule_conflict(lecturer, classroom, day, start_t, end_t, exclude_id=None):
    """Hoca veya sınıf çakışması var mı?"""
    qs = Schedule.objects.filter(day_of_week=day, start_time__lt=end_t, end_time__gt=start_t)
    if exclude_id:
        qs = qs.exclude(id=exclude_id)
    if lecturer and qs.filter(lecturer=lecturer).exists():
        return 'lecturer'
    if classroom and qs.filter(classroom=classroom).exists():
        return 'classroom'
    return None


def lecturer_is_unavailable(lecturer, day, start_t, end_t):
    """Hoca bu saate müsait değil mi?"""
    hour_str = f"{start_t.strftime('%H:%M')}-{end_t.strftime('%H:%M')}"
    return Unavailability.objects.filter(
        lecturer=lecturer, day=day, hour=hour_str
    ).exists()


def find_available_classroom(institution, day, start_t, end_t, preferred_type='LECTURE'):
    """Boş sınıf bul (tercih edilen tipte, yoksa herhangi biri)."""
    busy_ids = Schedule.objects.filter(
        day_of_week=day, start_time__lt=end_t, end_time__gt=start_t
    ).values_list('classroom_id', flat=True)
    available = Classroom.objects.filter(
        institution=institution
    ).exclude(id__in=busy_ids)

    preferred = available.filter(classroom_type=preferred_type).first()
    return preferred or available.first()


# ──────────────────────────── Custom JWT Login ──────────────────────────────

class MyTokenObtainPairSerializer(TokenObtainPairSerializer):
    def validate(self, attrs):
        username = attrs.get('username', '')

        # Hesap kilitli mi kontrol et
        try:
            user_obj = User.objects.get(username=username)
            if user_obj.locked_until and timezone.now() < user_obj.locked_until:
                remaining = max(1, int((user_obj.locked_until - timezone.now()).total_seconds() / 60))
                raise AuthenticationFailed(
                    f'Hesap geçici olarak kilitlendi. {remaining} dakika sonra tekrar deneyin.'
                )
        except User.DoesNotExist:
            pass

        try:
            data = super().validate(attrs)
        except Exception:
            # Başarısız giriş — sayacı artır
            try:
                user_obj = User.objects.get(username=username)
                user_obj.failed_login_attempts += 1
                if user_obj.failed_login_attempts >= 5:
                    user_obj.locked_until = timezone.now() + timedelta(minutes=15)
                    user_obj.failed_login_attempts = 0
                user_obj.save(update_fields=['failed_login_attempts', 'locked_until'])
                AuditLog.objects.create(
                    actor=user_obj,
                    action='LOGIN_FAIL',
                    model_name='User',
                    object_id=str(user_obj.pk),
                    description=f'Başarısız giriş denemesi: {username}',
                )
            except User.DoesNotExist:
                pass
            raise

        # Başarılı giriş — sayacı sıfırla
        self.user.failed_login_attempts = 0
        self.user.locked_until = None
        self.user.save(update_fields=['failed_login_attempts', 'locked_until'])
        AuditLog.objects.create(
            actor=self.user,
            action='LOGIN',
            model_name='User',
            object_id=str(self.user.pk),
            description=f'Başarılı giriş: {self.user.username}',
        )

        data['role'] = self.user.role
        data['username'] = self.user.username
        data['first_name'] = self.user.first_name
        data['last_name'] = self.user.last_name
        data['title'] = self.user.title
        data['must_change_password'] = self.user.must_change_password
        data['institution_name'] = self.user.institution.name if self.user.institution else 'Sistem Geneli'
        return data


class MyTokenObtainPairView(TokenObtainPairView):
    serializer_class = MyTokenObtainPairSerializer


# ──────────────────────────── ViewSets ──────────────────────────────────────

class InstitutionViewSet(viewsets.ModelViewSet):
    queryset = Institution.objects.all()
    serializer_class = InstitutionSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        user = self.request.user
        if not user.is_authenticated:
            return Institution.objects.none()
        if user.role == 'SUPERADMIN':
            return Institution.objects.all()
        if user.institution:
            return Institution.objects.filter(id=user.institution_id)
        return Institution.objects.none()


class DepartmentViewSet(viewsets.ModelViewSet):
    queryset = Department.objects.all()
    serializer_class = DepartmentSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        user = self.request.user
        if not user.is_authenticated:
            return Department.objects.none()
        if user.role == 'SUPERADMIN':
            return Department.objects.all()
        if user.institution:
            return Department.objects.filter(institution=user.institution)
        return Department.objects.none()


class UserViewSet(viewsets.ModelViewSet):
    queryset = User.objects.all()
    serializer_class = UserSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        user = self.request.user
        if not user.is_authenticated:
            return User.objects.none()
        if user.role == 'SUPERADMIN':
            qs = User.objects.all()
        elif user.role in ('ADMIN', 'STAFF', 'IT') and user.institution:
            qs = User.objects.filter(institution=user.institution).exclude(role='SUPERADMIN')
        else:
            qs = User.objects.filter(id=user.id)
        search = self.request.query_params.get('search')
        if search:
            qs = qs.filter(username__icontains=search)
        return qs

    @action(detail=False, methods=['post'], url_path='create-admin')
    def create_admin(self, request):
        """SUPERADMIN veya ADMIN herhangi bir kullanıcı oluşturabilir."""
        actor = request.user
        if actor.role not in ('SUPERADMIN', 'ADMIN', 'STAFF', 'IT'):
            return Response({'error': _t(request, 'Yetki yok.', 'You are not allowed to perform this action.')}, status=403)

        data = request.data
        role = data.get('role', 'LECTURER')

        # ADMIN yalnızca kendi kurumuna ve belirli rollere kullanıcı ekleyebilir
        if actor.role != 'SUPERADMIN':
            if role in ('SUPERADMIN',):
                return Response({'error': _t(request, 'Bu rolü sadece sistem yöneticisi atayabilir.', 'You are not allowed to create this role.')}, status=403)

        institution_id = data.get('institution_id') or (actor.institution_id if actor.role != 'SUPERADMIN' else None)
        if not institution_id:
            return Response({'error': _t(request, 'Kurum belirtilmedi.', 'Institution is required.')}, status=400)

        try:
            institution = Institution.objects.get(id=institution_id)
        except Institution.DoesNotExist:
            return Response({'error': _t(request, 'Kurum bulunamadı.', 'Institution not found.')}, status=404)

        username = data.get('username', '').strip()
        password = data.get('password', gen_password())
        first_name = data.get('first_name', '').strip()
        last_name = data.get('last_name', '').strip()
        title = data.get('title', '').strip()
        must_change = data.get('must_change_password', 'true') not in (False, 'false', 'False', '0')

        if not username:
            return Response({'error': _t(request, 'Kullanıcı adı zorunlu.', 'Username is required.')}, status=400)

        if User.objects.filter(username=username).exists():
            return Response({'error': _t(request, 'Bu kullanıcı adı zaten mevcut.', 'This username is already taken.')}, status=409)

        with transaction.atomic():
            user = User.objects.create(
                username=username,
                first_name=first_name,
                last_name=last_name,
                title=title,
                role=role,
                institution=institution,
                password=make_password(password),
                must_change_password=must_change,
            )
        audit(request.user, 'CREATE', 'User', user.pk,
              f'{role} hesabı oluşturuldu: {username} ({institution.name})', request=request)
        return Response({
            'message': f'{role} hesabı oluşturuldu.',
            'username': username,
            'password': password,
        }, status=201)

    @action(detail=False, methods=['patch'], url_path='update-profile')
    def update_profile(self, request):
        user = request.user
        first_name = request.data.get('first_name', '').strip()
        last_name = request.data.get('last_name', '').strip()
        email = request.data.get('email', '').strip()
        if first_name:
            user.first_name = first_name
        if last_name:
            user.last_name = last_name
        if email:
            user.email = email
        user.save(update_fields=['first_name', 'last_name', 'email'])
        return Response({
            'first_name': user.first_name,
            'last_name': user.last_name,
            'email': user.email,
        })

    @action(detail=False, methods=['post'], url_path='change-password')
    def change_password(self, request):
        user = request.user
        current = request.data.get('current_password', '')
        new_pw = request.data.get('new_password', '')

        if not current or not new_pw:
            return Response({'error': _t(request, 'Mevcut ve yeni şifre gerekli.', 'Current password and new password are required.')}, status=400)
        if not user.check_password(current):
            return Response({'error': _t(request, 'Mevcut şifre hatalı.', 'Current password is incorrect.')}, status=400)
        if len(new_pw) < 8:
            return Response({'error': _t(request, 'Yeni şifre en az 8 karakter olmalı.', 'New password must be at least 8 characters.')}, status=400)
        try:
            validate_password(new_pw, user)
        except DjangoValidationError as e:
            return Response({'error': ' '.join(e.messages)}, status=400)

        user.set_password(new_pw)
        user.must_change_password = False
        user.save(update_fields=['password', 'must_change_password'])
        audit(request.user, 'UPDATE', 'User', user.pk, f'Şifre değiştirildi: {user.username}', request=request)
        return Response({'message': 'Şifre güncellendi.'})


class ClassroomViewSet(viewsets.ModelViewSet):
    queryset = Classroom.objects.all()
    serializer_class = ClassroomSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        user = self.request.user
        if not user.is_authenticated:
            return Classroom.objects.none()
        if user.role == 'SUPERADMIN':
            return Classroom.objects.all()
        if user.institution:
            return Classroom.objects.filter(institution=user.institution)
        return Classroom.objects.none()

    def perform_create(self, serializer):
        user = self.request.user
        if user.role != 'SUPERADMIN' and user.institution:
            serializer.save(institution=user.institution)
        else:
            serializer.save()

    @action(detail=False, methods=['post'], url_path='bulk-import-excel',
            parser_classes=[MultiPartParser])
    @transaction.atomic
    def bulk_import_excel(self, request):
        user = request.user
        if user.role not in ('ADMIN', 'SUPERADMIN', 'STAFF', 'IT') or not user.institution:
            return Response({'error': 'Yetkisiz işlem.'}, status=403)

        file_obj = request.FILES.get('file')
        if not file_obj:
            return Response({'error': 'Dosya yok.'}, status=400)

        results = []
        content = file_obj.read()
        filename = file_obj.name.lower()

        try:
            if filename.endswith('.xlsx') or filename.endswith('.xls'):
                wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
                sheet = wb.active
                rows = [
                    [str(c).strip() if c is not None else '' for c in row]
                    for row in sheet.iter_rows(min_row=2, values_only=True)
                    if any(c for c in row)
                ]
            else:
                text = content.decode('utf-8', errors='replace')
                rows = []
                for line in text.splitlines():
                    line = line.strip()
                    if not line or line.startswith('#'):
                        continue
                    sep = '\t' if '\t' in line else ','
                    rows.append([c.strip() for c in line.split(sep)])

            for row in rows:
                if len(row) < 2:
                    continue
                room_code = row[0].upper()
                try:
                    cap = int(row[1])
                except ValueError:
                    cap = 30

                raw_type = row[2].upper() if len(row) >= 3 else ''
                TYPE_MAP = {
                    'LECTURE': 'LECTURE', 'DERSLİK': 'LECTURE', 'DERSLIK': 'LECTURE',
                    'LAB': 'LAB', 'LABORATORY': 'LAB', 'LABORATUVAR': 'LAB',
                    'COMPUTER_LAB': 'COMPUTER_LAB', 'BİLGİSAYAR': 'COMPUTER_LAB',
                    'BILGISAYAR': 'COMPUTER_LAB',
                    'SEMINAR': 'SEMINAR', 'SEMİNER': 'SEMINAR',
                }
                cls_type = TYPE_MAP.get(raw_type, 'LAB' if 'LAB' in room_code else 'LECTURE')

                obj, created = Classroom.objects.update_or_create(
                    room_code=room_code,
                    institution=user.institution,
                    defaults={'capacity': cap, 'classroom_type': cls_type}
                )
                results.append({
                    'room_code': room_code,
                    'capacity': str(cap),
                    'classroom_type': cls_type,
                    'created': str(created),
                })

            return Response(results, status=201)
        except Exception as e:
            return Response({'error': str(e)}, status=500)


class CourseViewSet(viewsets.ModelViewSet):
    queryset = Course.objects.all()
    serializer_class = CourseSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        user = self.request.user
        if not user.is_authenticated:
            return Course.objects.none()
        if user.role == 'SUPERADMIN':
            return Course.objects.all()
        if user.institution:
            return Course.objects.filter(department__institution=user.institution)
        return Course.objects.none()

    @action(detail=False, methods=['post'], url_path='bulk-import-excel',
            parser_classes=[MultiPartParser])
    @transaction.atomic
    def bulk_import_excel(self, request):
        user = request.user
        if user.role not in ('ADMIN', 'SUPERADMIN', 'STAFF', 'IT') or not user.institution:
            return Response({'error': 'Yetkisiz işlem.'}, status=403)

        file_obj = request.FILES.get('file')
        if not file_obj:
            return Response({'error': 'Dosya yok.'}, status=400)

        results = []
        content = file_obj.read()
        filename = file_obj.name.lower()

        try:
            if filename.endswith('.xlsx') or filename.endswith('.xls'):
                wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
                sheet = wb.active
                rows = [
                    [str(c).strip() if c is not None else '' for c in row]
                    for row in sheet.iter_rows(min_row=2, values_only=True)
                    if any(c for c in row)
                ]
            else:
                text = content.decode('utf-8', errors='replace')
                rows = []
                for line in text.splitlines():
                    line = line.strip()
                    if not line or line.startswith('#'):
                        continue
                    sep = '\t' if '\t' in line else ','
                    rows.append([c.strip() for c in line.split(sep)])

            for row in rows:
                if len(row) < 3:
                    continue

                if len(row) >= 4:
                    dept_name, course_code, course_name, lecturer_raw = row[0], row[1], row[2], row[3]
                else:
                    dept_name = 'Genel'
                    course_code, course_name, lecturer_raw = row[0], row[1], row[2]

                has_lab = False
                if len(row) >= 5:
                    has_lab = row[4].lower() in ('true', '1', 'evet', 'yes', 'var')

                department, _ = Department.objects.get_or_create(
                    name=dept_name, institution=user.institution
                )

                title, clean_name = extract_title_and_name(lecturer_raw)
                parts = clean_name.split()
                first_name = parts[0] if parts else clean_name
                last_name = ' '.join(parts[1:]) if len(parts) > 1 else ''

                username = make_username(first_name, last_name)
                password = gen_password()

                lecturer, created = User.objects.get_or_create(
                    username=username,
                    defaults={
                        'first_name': first_name,
                        'last_name': last_name,
                        'title': title,
                        'role': 'LECTURER',
                        'institution': user.institution,
                        'password': make_password(password),
                        'must_change_password': True,
                    }
                )

                course, _ = Course.objects.update_or_create(
                    course_code=course_code,
                    department=department,
                    defaults={
                        'course_name': course_name,
                        'has_lab': has_lab,
                    }
                )

                results.append({
                    'course': course_code,
                    'lecturer': lecturer_raw,
                    'generated_user': username,
                    'generated_pass': password if created else '••••••',
                })

            return Response(results, status=201)
        except Exception as e:
            return Response({'error': str(e)}, status=500)


class ScheduleViewSet(viewsets.ModelViewSet):
    queryset = Schedule.objects.select_related(
        'course', 'lecturer', 'classroom'
    ).all()
    serializer_class = ScheduleSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        user = self.request.user
        if not user.is_authenticated:
            return Schedule.objects.none()

        qs = Schedule.objects.select_related('course', 'lecturer', 'classroom').filter(is_deleted=False)

        if user.role == 'SUPERADMIN':
            return qs
        if user.role in ('ADMIN', 'STAFF', 'IT') and user.institution:
            return qs.filter(course__department__institution=user.institution)
        if user.role == 'LECTURER':
            return qs.filter(lecturer=user)
        return qs.none()

    def destroy(self, request, *args, **kwargs):
        schedule = self.get_object()
        schedule.is_deleted = True
        schedule.deleted_at = timezone.now()
        schedule.save(update_fields=['is_deleted', 'deleted_at'])
        audit(request.user, 'DELETE', 'Schedule', schedule.pk,
              f'Program silindi: {schedule.course.course_code} {schedule.day_of_week} {schedule.start_time}', request=request)
        return Response(status=status.HTTP_204_NO_CONTENT)

    def create(self, request, *args, **kwargs):
        data = request.data
        user = request.user

        # day_of_week normalize
        day = normalize_day(data.get('day_of_week', ''))

        try:
            start_t = time_type.fromisoformat(str(data.get('start_time', '')))
            end_t = time_type.fromisoformat(str(data.get('end_time', '')))
        except ValueError:
            return Response({'error': 'Geçersiz saat formatı. HH:MM kullanın.'}, status=400)

        lecturer_id = data.get('lecturer')
        classroom_id = data.get('classroom')
        course_id = data.get('course')

        if not course_id:
            return Response({'error': 'Ders (course) zorunlu.'}, status=400)

        lecturer = None
        if lecturer_id:
            try:
                lecturer = User.objects.get(id=lecturer_id)
            except User.DoesNotExist:
                return Response({'error': 'Hoca bulunamadı.'}, status=404)

        classroom = None
        if classroom_id:
            try:
                classroom = Classroom.objects.get(id=classroom_id)
            except Classroom.DoesNotExist:
                return Response({'error': 'Sınıf bulunamadı.'}, status=404)

        # Çakışma kontrolü
        conflict = has_schedule_conflict(lecturer, classroom, day, start_t, end_t)
        if conflict == 'lecturer':
            return Response(
                {'error': _t(request,
                    f'⚠️ Çakışma: {lecturer.get_full_name() or lecturer.username} bu saatte başka bir derste.',
                    f'Lecturer {lecturer.get_full_name() or lecturer.username} already has a class at this time.')},
                status=409
            )
        if conflict == 'classroom':
            return Response(
                {'error': _t(request,
                    f'⚠️ Çakışma: {classroom.room_code} bu saatte dolu.',
                    f'Classroom {classroom.room_code} is already booked at this time.')},
                status=409
            )

        try:
            course = Course.objects.get(id=course_id)
        except Course.DoesNotExist:
            return Response({'error': 'Ders bulunamadı.'}, status=404)

        schedule = Schedule.objects.create(
            course=course,
            lecturer=lecturer,
            classroom=classroom,
            day_of_week=day,
            start_time=start_t,
            end_time=end_t,
            session_type=data.get('session_type', 'LECTURE'),
        )
        return Response(ScheduleSerializer(schedule).data, status=201)

    @action(detail=False, methods=['get'], url_path='admin-summary')
    def admin_summary(self, request):
        user = request.user
        if user.role not in ('ADMIN', 'SUPERADMIN', 'STAFF', 'IT'):
            return Response({'error': 'Yetki yok.'}, status=403)
        if not user.institution:
            return Response({'error': 'Kurum bilgisi yok.'}, status=400)

        inst = user.institution

        # Atanmamış hocalar (hiç schedule'ı olmayan LECTURER/STAFF)
        scheduled_lec_ids = Schedule.objects.filter(
            course__department__institution=inst
        ).values_list('lecturer_id', flat=True)
        unassigned_lecturers = User.objects.filter(
            institution=inst, role__in=('LECTURER', 'STAFF')
        ).exclude(id__in=scheduled_lec_ids).values(
            'id', 'username', 'first_name', 'last_name', 'title'
        )

        # Atanmamış dersler (hiç schedule'ı olmayan dersler)
        scheduled_course_ids = Schedule.objects.filter(
            course__department__institution=inst
        ).values_list('course_id', flat=True)
        unassigned_courses = Course.objects.filter(
            department__institution=inst
        ).exclude(id__in=scheduled_course_ids).values(
            'id', 'course_code', 'course_name'
        )

        # Müsait sınıflar (bugünkü program saatlerinde boş olanlar — tümünü listele)
        all_busy_ids = Schedule.objects.filter(
            course__department__institution=inst
        ).values_list('classroom_id', flat=True)
        available_classrooms = Classroom.objects.filter(
            institution=inst
        ).exclude(id__in=all_busy_ids).values(
            'id', 'room_code', 'capacity', 'classroom_type'
        )

        return Response({
            'unassigned_lecturers': list(unassigned_lecturers),
            'unassigned_courses': list(unassigned_courses),
            'available_classrooms': list(available_classrooms),
        })

    @action(detail=False, methods=['get'])
    def available_for_slot(self, request):
        """
        Returns classrooms available for a given time slot.
        Query params: day, start_time, end_time, session_type (LECTURE or LAB)
        Optional: course_id (to auto-determine required classroom types)
        """
        day          = request.query_params.get('day', '')
        start_time   = request.query_params.get('start_time', '')
        end_time     = request.query_params.get('end_time', '')
        session_type = request.query_params.get('session_type', 'LECTURE')
        course_id    = request.query_params.get('course_id')

        if not all([day, start_time, end_time]):
            return Response(
                {'error': _t(request, 'day, start_time ve end_time zorunlu.', 'day, start_time and end_time are required.')},
                status=400
            )

        # Determine which classroom types are appropriate for this session type
        if session_type == 'LAB':
            allowed_types = ['LAB', 'COMPUTER_LAB']
        else:
            allowed_types = ['LECTURE', 'SEMINAR', 'OTHER']

        # If course_id given, refine based on course's has_lab flag
        if course_id:
            try:
                course = Course.objects.get(pk=course_id)
                if not course.has_lab:
                    # Non-lab course: any classroom type is fine
                    allowed_types = ['LECTURE', 'LAB', 'COMPUTER_LAB', 'SEMINAR', 'OTHER']
            except Course.DoesNotExist:
                pass

        institution = request.user.institution
        if not institution:
            return Response({'error': _t(request, 'Kurum bilgisi yok.', 'Institution not found.')}, status=400)

        classrooms = Classroom.objects.filter(institution=institution, classroom_type__in=allowed_types)

        try:
            start_t = time_type.fromisoformat(start_time)
            end_t = time_type.fromisoformat(end_time)
        except ValueError:
            return Response({'error': _t(request, 'Geçersiz saat formatı. HH:MM kullanın.', 'Invalid time format. Use HH:MM.')}, status=400)

        booked_ids = Schedule.objects.filter(
            classroom__institution=institution,
            day_of_week=day,
            start_time=start_t,
            end_time=end_t
        ).values_list('classroom_id', flat=True)

        available = classrooms.exclude(id__in=booked_ids)
        serializer = ClassroomSerializer(available, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['post'], url_path='generate-auto')
    @transaction.atomic
    def generate_auto(self, request):
        """Greedy otomatik program oluşturucu."""
        user = request.user
        if user.role not in ('ADMIN', 'SUPERADMIN', 'STAFF', 'IT'):
            return Response({'error': 'Yetki yok.'}, status=403)
        if not user.institution:
            return Response({'error': 'Kurum bilgisi yok.'}, status=400)

        inst = user.institution
        scheduled_count = 0
        skipped = []

        # Atanmamış dersler
        scheduled_ids = Schedule.objects.filter(
            course__department__institution=inst
        ).values_list('course_id', flat=True)
        courses = Course.objects.filter(
            department__institution=inst
        ).exclude(id__in=scheduled_ids)

        for course in courses:
            # Bu kurs için kullanılabilecek hocayı bul
            # Önce daha önce atanmış bir hocanın var mı kontrol et
            existing = Schedule.objects.filter(course=course).first()
            if existing and existing.lecturer:
                lecturer = existing.lecturer
            else:
                # LECTURER rolündeki kurumun ilk uygun hocasını bul
                lecturer = User.objects.filter(
                    institution=inst, role__in=('LECTURER', 'STAFF')
                ).first()

            if not lecturer:
                skipped.append(f"{course.course_code}: Uygun hoca yok")
                continue

            assigned = False
            for day, start_t, end_t in ALL_SLOTS:
                # Hoca müsaitliği
                if lecturer_is_unavailable(lecturer, day, start_t, end_t):
                    continue
                # Çakışma
                conflict = has_schedule_conflict(lecturer, None, day, start_t, end_t)
                if conflict:
                    continue
                # Uygun sınıf
                classroom = find_available_classroom(inst, day, start_t, end_t, 'LECTURE')
                if not classroom:
                    continue

                Schedule.objects.create(
                    course=course,
                    lecturer=lecturer,
                    classroom=classroom,
                    day_of_week=day,
                    start_time=start_t,
                    end_time=end_t,
                    session_type='LECTURE',
                )
                scheduled_count += 1
                assigned = True
                break

            # Lab bileşeni de varsa ayrı slot ata
            if assigned and course.has_lab:
                for day, start_t, end_t in ALL_SLOTS:
                    if lecturer_is_unavailable(lecturer, day, start_t, end_t):
                        continue
                    conflict = has_schedule_conflict(lecturer, None, day, start_t, end_t)
                    if conflict:
                        continue
                    lab_room = find_available_classroom(inst, day, start_t, end_t, 'LAB')
                    if not lab_room:
                        lab_room = find_available_classroom(inst, day, start_t, end_t, 'COMPUTER_LAB')
                    if not lab_room:
                        continue

                    Schedule.objects.create(
                        course=course,
                        lecturer=lecturer,
                        classroom=lab_room,
                        day_of_week=day,
                        start_time=start_t,
                        end_time=end_t,
                        session_type='LAB',
                    )
                    scheduled_count += 1
                    break

            if not assigned:
                skipped.append(f"{course.course_code}: Uygun slot bulunamadı")

        return Response({
            'scheduled_count': scheduled_count,
            'skipped': skipped,
            'message': f'{scheduled_count} ders otomatik programlandı.',
        })


class UnavailabilityViewSet(viewsets.ViewSet):
    permission_classes = [IsAuthenticated]

    def list(self, request):
        """GET /api/unavailability/ — mevcut kullanıcının meşgul slotları"""
        if not request.user.is_authenticated:
            return Response([], status=200)
        data = Unavailability.objects.filter(
            lecturer=request.user
        ).values('day', 'hour')
        return Response(list(data))

    @action(detail=False, methods=['post'])
    def sync(self, request):
        """POST /api/unavailability/sync/ — tümünü sil, yeniden yaz"""
        if not request.user.is_authenticated:
            return Response({'error': 'Oturum açılı değil.'}, status=401)
        Unavailability.objects.filter(lecturer=request.user).delete()
        to_create = []
        for slot in request.data:
            day = slot.get('day', '').strip()
            hour = slot.get('hour', '').strip()
            if day and hour:
                to_create.append(
                    Unavailability(lecturer=request.user, day=day, hour=hour)
                )
        Unavailability.objects.bulk_create(to_create, ignore_conflicts=True)
        return Response({'message': f'{len(to_create)} tercih kaydedildi.'})


class StudentEnrollmentViewSet(viewsets.ModelViewSet):
    serializer_class = StudentEnrollmentSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        user = self.request.user
        if user.role in ('SUPERADMIN', 'ADMIN'):
            qs = StudentEnrollment.objects.select_related('student', 'course', 'course__department')
            if user.role == 'ADMIN':
                qs = qs.filter(course__department__institution=user.institution)
            return qs
        if user.role == 'STUDENT':
            return StudentEnrollment.objects.filter(student=user).select_related('course', 'course__department')
        return StudentEnrollment.objects.none()

    def perform_create(self, serializer):
        user = self.request.user
        # ADMIN/SUPERADMIN can enroll any student; STUDENT can self-enroll
        if user.role == 'STUDENT':
            serializer.save(student=user)
        else:
            serializer.save()

    @action(detail=False, methods=['get'])
    def my_schedule(self, request):
        """Returns the full schedule for the authenticated student's enrolled courses."""
        if request.user.role != 'STUDENT':
            return Response({'error': 'Only students can access this.'}, status=403)
        enrolled_course_ids = StudentEnrollment.objects.filter(
            student=request.user
        ).values_list('course_id', flat=True)
        schedules = Schedule.objects.filter(
            course_id__in=enrolled_course_ids
        ).select_related('course', 'lecturer', 'classroom')
        serializer = ScheduleSerializer(schedules, many=True)
        return Response(serializer.data)


class AnnouncementViewSet(viewsets.ModelViewSet):
    serializer_class   = AnnouncementSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        user = self.request.user
        if user.role == 'SUPERADMIN':
            qs = Announcement.objects.filter(is_active=True)
        else:
            qs = Announcement.objects.filter(institution=user.institution, is_active=True)
        if user.role in ('ADMIN', 'SUPERADMIN'):
            return qs
        return qs.filter(audience__in=['ALL', user.role])

    def perform_create(self, serializer):
        serializer.save(
            created_by=self.request.user,
            institution=self.request.user.institution
        )

    @action(detail=True, methods=['post'])
    def mark_read(self, request, pk=None):
        """Mark a specific announcement as read for the current user."""
        announcement = self.get_object()
        notif, _ = UserNotification.objects.get_or_create(
            user=request.user, announcement=announcement
        )
        if not notif.is_read:
            notif.is_read = True
            notif.read_at = timezone.now()
            notif.save()
        return Response({'status': 'ok'})

    @action(detail=False, methods=['post'])
    def mark_all_read(self, request):
        """Mark all announcements as read for the current user."""
        announcements = self.get_queryset()
        for ann in announcements:
            UserNotification.objects.update_or_create(
                user=request.user,
                announcement=ann,
                defaults={'is_read': True, 'read_at': timezone.now()}
            )
        return Response({'status': 'ok'})

    @action(detail=False, methods=['get'])
    def unread_count(self, request):
        """Returns count of unread announcements for the current user."""
        announcements = self.get_queryset()
        read_ids = UserNotification.objects.filter(
            user=request.user, is_read=True
        ).values_list('announcement_id', flat=True)
        count = announcements.exclude(id__in=read_ids).count()
        return Response({'unread_count': count})


class FCMTokenViewSet(viewsets.ViewSet):
    """Allows mobile clients to register/update their FCM push token."""
    permission_classes = [IsAuthenticated]

    @action(detail=False, methods=['post'])
    def register(self, request):
        token = request.data.get('fcm_token', '').strip()
        if not token:
            return Response({'error': _t(request, 'Token gerekli.', 'Token is required.')}, status=400)
        request.user.fcm_token = token
        request.user.save(update_fields=['fcm_token'])
        return Response({'status': 'registered'})


class AttendanceViewSet(viewsets.ViewSet):
    permission_classes = [IsAuthenticated]

    @action(detail=False, methods=['post'])
    def create_session(self, request):
        """LECTURER creates an attendance session for a schedule slot. Token valid 90 seconds."""
        if request.user.role not in ('LECTURER', 'ADMIN', 'SUPERADMIN'):
            return Response({'error': _t(request, 'Sadece öğretim üyeleri yoklama başlatabilir.', 'Only lecturers can start attendance.')}, status=403)

        schedule_id = request.data.get('schedule_id')
        session_date = request.data.get('session_date')  # YYYY-MM-DD

        if not schedule_id or not session_date:
            return Response({'error': _t(request, 'schedule_id ve session_date zorunlu.', 'schedule_id and session_date are required.')}, status=400)

        try:
            schedule = Schedule.objects.get(pk=schedule_id)
        except Schedule.DoesNotExist:
            return Response({'error': _t(request, 'Ders bulunamadı.', 'Schedule not found.')}, status=404)

        # Deactivate any existing active sessions for this schedule+date
        AttendanceSession.objects.filter(
            schedule=schedule, session_date=session_date, is_active=True
        ).update(is_active=False)

        expires_at = timezone.now() + timedelta(seconds=90)
        session = AttendanceSession.objects.create(
            schedule=schedule,
            created_by=request.user,
            expires_at=expires_at,
            session_date=session_date
        )
        serializer = AttendanceSessionSerializer(session)
        return Response(serializer.data, status=201)

    @action(detail=False, methods=['post'])
    def check_in(self, request):
        """STUDENT checks in by submitting the QR token."""
        if request.user.role != 'STUDENT':
            return Response({'error': _t(request, 'Sadece öğrenciler yoklamaya katılabilir.', 'Only students can check in.')}, status=403)

        token = request.data.get('token', '').strip()
        if not token:
            return Response({'error': _t(request, 'Token gerekli.', 'Token is required.')}, status=400)

        try:
            session = AttendanceSession.objects.get(token=token, is_active=True)
        except AttendanceSession.DoesNotExist:
            return Response({'error': _t(request, 'Geçersiz veya süresi dolmuş token.', 'Invalid or expired token.')}, status=404)

        if session.is_expired():
            return Response({'error': _t(request, 'QR kodunun süresi doldu.', 'QR code has expired.')}, status=410)

        # Check student is enrolled in this course
        enrolled = StudentEnrollment.objects.filter(
            student=request.user, course=session.schedule.course
        ).exists()
        if not enrolled:
            return Response({'error': _t(request, 'Bu derse kayıtlı değilsiniz.', 'You are not enrolled in this course.')}, status=403)

        record, created = AttendanceRecord.objects.get_or_create(
            session=session, student=request.user
        )
        if not created:
            return Response({'error': _t(request, 'Zaten yoklamaya katıldınız.', 'You have already checked in.')}, status=409)

        return Response({'status': 'checked_in', 'course': session.schedule.course.course_code}, status=201)

    @action(detail=False, methods=['get'])
    def session_records(self, request):
        """LECTURER views attendance records for a session."""
        session_id = request.query_params.get('session_id')
        if not session_id:
            return Response({'error': 'session_id required'}, status=400)
        try:
            session = AttendanceSession.objects.get(pk=session_id)
        except AttendanceSession.DoesNotExist:
            return Response({'error': 'Not found'}, status=404)
        records = session.records.select_related('student')
        serializer = AttendanceRecordSerializer(records, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['get'])
    def my_sessions(self, request):
        """LECTURER gets their attendance sessions."""
        if request.user.role not in ('LECTURER', 'ADMIN', 'SUPERADMIN'):
            return Response({'error': 'Forbidden'}, status=403)
        sessions = AttendanceSession.objects.filter(
            created_by=request.user
        ).select_related('schedule__course')[:50]
        serializer = AttendanceSessionSerializer(sessions, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['get'])
    def my_attendance(self, request):
        """STUDENT views their own attendance records."""
        if request.user.role != 'STUDENT':
            return Response({'error': 'Forbidden'}, status=403)
        records = AttendanceRecord.objects.filter(
            student=request.user
        ).select_related('session__schedule__course')[:100]
        serializer = AttendanceRecordSerializer(records, many=True)
        return Response(serializer.data)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_schedule_pdf(request):
    """
    Exports the weekly schedule as a PDF.
    Query params:
      - type: 'institution' (all schedules, admin only) | 'lecturer' (for a specific lecturer) | 'course' (for a specific course)
      - lecturer_id: required if type=lecturer
      - course_id: required if type=course
    """
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    from io import BytesIO

    export_type = request.query_params.get('type', 'institution')
    user = request.user

    # Build queryset
    qs = Schedule.objects.select_related('course', 'lecturer', 'classroom').filter(
        course__department__institution=user.institution
    )

    if export_type == 'lecturer':
        lecturer_id = request.query_params.get('lecturer_id', user.id)
        qs = qs.filter(lecturer_id=lecturer_id)
        title_text = "Öğretim Üyesi Ders Programı"
    elif export_type == 'course':
        course_id = request.query_params.get('course_id')
        if course_id:
            qs = qs.filter(course_id=course_id)
        title_text = "Ders Programı"
    else:
        if user.role not in ('ADMIN', 'SUPERADMIN'):
            # Non-admin: show only their own schedule
            qs = qs.filter(lecturer=user)
        title_text = f"{user.institution.name} – Haftalık Ders Programı"

    DAY_ORDER = ['Pazartesi', 'Salı', 'Çarşamba', 'Perşembe', 'Cuma']
    schedules = list(qs)
    schedules.sort(key=lambda s: (
        DAY_ORDER.index(s.day_of_week) if s.day_of_week in DAY_ORDER else 99,
        str(s.start_time)
    ))

    # Build PDF
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        rightMargin=1.5*cm, leftMargin=1.5*cm,
        topMargin=1.5*cm, bottomMargin=1.5*cm
    )

    styles = getSampleStyleSheet()
    elements = []

    # Title
    title_style = styles['Title']
    elements.append(Paragraph(title_text, title_style))
    elements.append(Spacer(1, 0.4*cm))

    # Table header
    header = ['Gün', 'Başlangıç', 'Bitiş', 'Ders Kodu', 'Ders Adı', 'Tip', 'Öğretim Üyesi', 'Sınıf']
    table_data = [header]

    SESSION_LABELS = {'LECTURE': 'Teorik', 'LAB': 'Lab'}
    for s in schedules:
        lecturer_name = ''
        if s.lecturer:
            lecturer_name = f"{s.lecturer.first_name} {s.lecturer.last_name}".strip() or s.lecturer.username
        table_data.append([
            s.day_of_week or '',
            str(s.start_time)[:5] if s.start_time else '',
            str(s.end_time)[:5] if s.end_time else '',
            s.course.course_code if s.course else '',
            s.course.course_name if s.course else '',
            SESSION_LABELS.get(s.session_type, s.session_type or ''),
            lecturer_name,
            s.classroom.room_code if s.classroom else '',
        ])

    if len(table_data) == 1:
        table_data.append(['Program bulunamadı', '', '', '', '', '', '', ''])

    col_widths = [3*cm, 2*cm, 2*cm, 2.5*cm, 6*cm, 2*cm, 5*cm, 2.5*cm]
    table = Table(table_data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND',   (0, 0), (-1, 0), colors.HexColor('#1565C0')),
        ('TEXTCOLOR',    (0, 0), (-1, 0), colors.white),
        ('FONTNAME',     (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE',     (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING',(0, 0), (-1, 0), 8),
        ('TOPPADDING',   (0, 0), (-1, 0), 8),
        ('ALIGN',        (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN',       (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTNAME',     (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE',     (0, 1), (-1, -1), 9),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#EEF2F8')]),
        ('GRID',         (0, 0), (-1, -1), 0.5, colors.HexColor('#CCCCCC')),
        ('LEFTPADDING',  (0, 0), (-1, -1), 6),
        ('RIGHTPADDING', (0, 0), (-1, -1), 6),
    ]))
    elements.append(table)

    doc.build(elements)
    buffer.seek(0)

    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="schedule.pdf"'
    return response


class ChatViewSet(viewsets.ViewSet):
    permission_classes = [IsAuthenticated]

    @action(detail=False, methods=['get'])
    def inbox(self, request):
        """Returns a list of conversations (unique users the current user has chatted with)."""
        user = request.user
        from django.db.models import Q, Max, Subquery, OuterRef
        partners_sent = ChatMessage.objects.filter(sender=user).values_list('receiver_id', flat=True)
        partners_recv = ChatMessage.objects.filter(receiver=user).values_list('sender_id', flat=True)
        partner_ids = set(list(partners_sent) + list(partners_recv))

        conversations = []
        for pid in partner_ids:
            try:
                partner = User.objects.get(pk=pid)
            except User.DoesNotExist:
                continue
            last_msg = ChatMessage.objects.filter(
                Q(sender=user, receiver=partner) | Q(sender=partner, receiver=user)
            ).order_by('-created_at').first()
            unread = ChatMessage.objects.filter(sender=partner, receiver=user, is_read=False).count()
            if last_msg:
                conversations.append({
                    'partner_id': partner.id,
                    'partner_username': partner.username,
                    'partner_name': f"{partner.first_name} {partner.last_name}".strip() or partner.username,
                    'partner_role': partner.role,
                    'last_message': last_msg.content[:80],
                    'last_message_at': last_msg.created_at.isoformat(),
                    'unread_count': unread,
                    'is_mine': last_msg.sender_id == user.id,
                })
        conversations.sort(key=lambda x: x['last_message_at'], reverse=True)
        return Response(conversations)

    @action(detail=False, methods=['get'])
    def messages(self, request):
        """Returns messages between current user and a partner."""
        from django.db.models import Q
        partner_id = request.query_params.get('partner_id')
        if not partner_id:
            return Response({'error': _t(request, 'partner_id gerekli.', 'partner_id is required.')}, status=400)
        try:
            partner = User.objects.get(pk=partner_id)
        except User.DoesNotExist:
            return Response({'error': _t(request, 'Kullanıcı bulunamadı.', 'User not found.')}, status=404)
        msgs = ChatMessage.objects.filter(
            Q(sender=request.user, receiver=partner) | Q(sender=partner, receiver=request.user)
        ).order_by('created_at')
        # Mark received messages as read
        msgs.filter(receiver=request.user, is_read=False).update(is_read=True, read_at=timezone.now())
        serializer = ChatMessageSerializer(msgs, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['post'])
    def send(self, request):
        """Send a message to another user."""
        from django.db.models import Q
        receiver_id = request.data.get('receiver_id')
        content = request.data.get('content', '').strip()
        if not receiver_id or not content:
            return Response({'error': _t(request, 'receiver_id ve content zorunlu.', 'receiver_id and content are required.')}, status=400)
        try:
            receiver = User.objects.get(pk=receiver_id, institution=request.user.institution)
        except User.DoesNotExist:
            return Response({'error': _t(request, 'Kullanıcı bulunamadı.', 'User not found.')}, status=404)
        msg = ChatMessage.objects.create(sender=request.user, receiver=receiver, content=content)
        serializer = ChatMessageSerializer(msg)
        return Response(serializer.data, status=201)

    @action(detail=False, methods=['get'])
    def contacts(self, request):
        """Returns users in the same institution the current user can chat with."""
        users = User.objects.filter(
            institution=request.user.institution, is_active=True
        ).exclude(pk=request.user.pk).order_by('first_name', 'last_name')
        data = [{
            'id': u.id,
            'username': u.username,
            'name': f"{u.first_name} {u.last_name}".strip() or u.username,
            'role': u.role,
        } for u in users]
        return Response(data)

    @action(detail=False, methods=['get'])
    def unread_count(self, request):
        count = ChatMessage.objects.filter(receiver=request.user, is_read=False).count()
        return Response({'unread_count': count})


class CourseMaterialViewSet(viewsets.ModelViewSet):
    serializer_class   = CourseMaterialSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        user = self.request.user
        course_id = self.request.query_params.get('course_id')
        qs = CourseMaterial.objects.select_related('course', 'uploaded_by')
        if user.role == 'STUDENT':
            enrolled_ids = StudentEnrollment.objects.filter(student=user).values_list('course_id', flat=True)
            qs = qs.filter(course_id__in=enrolled_ids)
        elif user.role == 'LECTURER':
            qs = qs.filter(course__department__institution=user.institution)
        else:
            qs = qs.filter(course__department__institution=user.institution)
        if course_id:
            qs = qs.filter(course_id=course_id)
        return qs

    def perform_create(self, serializer):
        serializer.save(uploaded_by=self.request.user)

    def create(self, request, *args, **kwargs):
        if request.user.role not in ('LECTURER', 'ADMIN', 'SUPERADMIN'):
            return Response({'error': _t(request, 'Sadece öğretim üyeleri materyal yükleyebilir.', 'Only lecturers can upload materials.')}, status=403)

        file = request.FILES.get('file')
        if file:
            max_size = 50 * 1024 * 1024  # 50 MB
            if file.size > max_size:
                return Response({'error': _t(request, 'Dosya boyutu 50MB sınırını aşamaz.', 'File size cannot exceed 50MB.')}, status=400)
            allowed_types = {
                'application/pdf', 'image/jpeg', 'image/png', 'image/gif',
                'application/msword',
                'application/vnd.openxmlformats-officedocument.wordprocessingml.document',
                'application/vnd.ms-powerpoint',
                'application/vnd.openxmlformats-officedocument.presentationml.presentation',
                'application/vnd.ms-excel',
                'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                'text/plain', 'application/zip', 'application/x-zip-compressed',
            }
            if file.content_type not in allowed_types:
                return Response({'error': _t(request, f'Desteklenmeyen dosya türü: {file.content_type}', f'Unsupported file type: {file.content_type}')}, status=400)

        return super().create(request, *args, **kwargs)


class GradeViewSet(viewsets.ModelViewSet):
    serializer_class   = GradeSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        user = self.request.user
        course_id = self.request.query_params.get('course_id')
        if user.role == 'STUDENT':
            qs = Grade.objects.filter(student=user).select_related('course', 'graded_by')
        elif user.role in ('LECTURER', 'ADMIN', 'SUPERADMIN'):
            qs = Grade.objects.filter(course__department__institution=user.institution).select_related('student', 'course')
        else:
            qs = Grade.objects.none()
        if course_id:
            qs = qs.filter(course_id=course_id)
        return qs

    def perform_create(self, serializer):
        serializer.save(graded_by=self.request.user)

    @action(detail=False, methods=['get'])
    def my_grades(self, request):
        """Student: get all their grades with averages per course."""
        if request.user.role != 'STUDENT':
            return Response({'error': 'Only students.'}, status=403)
        grades = Grade.objects.filter(student=request.user).select_related('course')
        serializer = GradeSerializer(grades, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['get'])
    def course_grades(self, request):
        """Lecturer: get all grades for a course."""
        course_id = request.query_params.get('course_id')
        if not course_id:
            return Response({'error': 'course_id required'}, status=400)
        if request.user.role not in ('LECTURER', 'ADMIN', 'SUPERADMIN'):
            return Response({'error': 'Forbidden'}, status=403)
        grades = Grade.objects.filter(course_id=course_id).select_related('student')
        serializer = GradeSerializer(grades, many=True)
        return Response(serializer.data)
